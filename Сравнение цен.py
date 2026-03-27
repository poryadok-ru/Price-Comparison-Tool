import openpyxl
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def split_multiple_barcodes(barcode):
    """Разделяет строку с несколькими штрихкодами на отдельные значения"""
    if barcode is None:
        return []
    barcode_str = str(barcode).strip()
    if not barcode_str:
        return []
    for sep in ['\n', '\r', '\t', ',', ';']:
        barcode_str = barcode_str.replace(sep, ' ')
    # Убираем лишние символы: дефисы в начале, пробелы, неразрывные пробелы
    results = []
    for x in barcode_str.split(' '):
        cleaned = x.strip().strip('\xa0').lstrip('-')
        if cleaned:
            results.append(cleaned)
    return results


def to_float(value):
    """Преобразует значение в число, обрабатывая разные форматы"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).strip().replace(' ', '').replace('\xa0', '').replace(',', '.')
        return float(cleaned)
    except:
        return None


def log(message):
    """Логирование для отладки"""
    print(message)
    with open('price_update_log.txt', 'a', encoding='utf-8') as f:
        f.write(message + '\n')


# Очистка лога
with open('price_update_log.txt', 'w', encoding='utf-8') as f:
    f.write('')

YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# ============================================================
# НАСТРОЙКИ ПУТЕЙ — измените под свою структуру
# ============================================================
template_path = r'\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\В.Гречушкин\Общая\Доработки\Сравнение цен\шаблон.xlsx'
orders_dir = r'\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\В.Гречушкин\Общая\Доработки\Сравнение цен\Заказы'

log("=" * 60)
log("НАЧАЛО ОБРАБОТКИ")
log("=" * 60)

# ============================================================
# ШАГ 0: Разъединить объединённые ячейки в файлах заказов
# ============================================================
files = [f for f in os.listdir(orders_dir) if f.endswith('.xlsx')]
log(f"\nФайлов заказов найдено: {len(files)}")

for file_name in files:
    file_path = os.path.join(orders_dir, file_name)
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        merged_cell_ranges = list(sheet.merged_cells.ranges)
        if merged_cell_ranges:
            for merged_cell_range in merged_cell_ranges:
                sheet.unmerge_cells(str(merged_cell_range))
            workbook.save(file_path)
            log(f"  Разъединено {len(merged_cell_ranges)} ячеек в {file_name}")
        workbook.close()
    except Exception as e:
        log(f"  ОШИБКА при разъединении в {file_name}: {e}")

# ============================================================
# ШАГ 1: Загрузка шаблона
# ============================================================
try:
    template_wb = load_workbook(template_path)
    template_ws = template_wb.active
    log(f"\nШаблон загружен: {template_ws.max_row} строк, {template_ws.max_column} столбцов")
except Exception as e:
    log(f"ОШИБКА при загрузке шаблона: {e}")
    exit()

# Определяем индексы столбцов по заголовкам (ищем только в первых 20 столбцах)
col_map = {}
for col_idx in range(1, min(template_ws.max_column + 1, 50)):
    header = template_ws.cell(row=1, column=col_idx).value
    if header and str(header).strip():
        col_map[str(header).strip()] = col_idx

log(f"Найденные столбцы: {col_map}")

price_agreed_col = col_map.get("Цена согласованная")
price_actual_col = col_map.get("Цена фактическая")
price_diff_col = col_map.get("Разница цен")
code_col = col_map.get("Код")
barcode_col = col_map.get("Штрихкод")

if not all([price_agreed_col, price_actual_col, price_diff_col]):
    log(f"ОШИБКА: Не найдены столбцы цен! agreed={price_agreed_col}, actual={price_actual_col}, diff={price_diff_col}")
    exit()

log(f"Столбцы: Код={code_col}, Штрихкод={barcode_col}, Согласованная={price_agreed_col}, Фактическая={price_actual_col}, Разница={price_diff_col}")

# ============================================================
# ШАГ 2: Сбор поисковых значений из шаблона
# ============================================================
search_values = []
for row_idx in range(2, template_ws.max_row + 1):
    all_values = set()

    if code_col:
        code = template_ws.cell(row=row_idx, column=code_col).value
        if code is not None and str(code).strip():
            all_values.add(str(code).strip())

    if barcode_col:
        barcode = template_ws.cell(row=row_idx, column=barcode_col).value
        if barcode is not None:
            for bc in split_multiple_barcodes(barcode):
                all_values.add(bc)

    if all_values:
        search_values.append((all_values, row_idx))

log(f"\nСтрок для поиска: {len(search_values)}")
log(f"Примеры поисковых значений (первые 5):")
for vals, row in search_values[:5]:
    log(f"  Строка {row}: {vals}")

# ============================================================
# ШАГ 3: ДИАГНОСТИКА файлов заказов — определяем структуру
# ============================================================
log(f"\n{'=' * 60}")
log("ДИАГНОСТИКА ФАЙЛОВ ЗАКАЗОВ")
log("=" * 60)

for filename in sorted(os.listdir(orders_dir)):
    if not filename.endswith('.xlsx') or filename == 'шаблон.xlsx':
        continue
    file_path = os.path.join(orders_dir, filename)
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        log(f"\n--- {filename} ---")
        log(f"  Размер: {ws.max_row} строк x {ws.max_column} столбцов")

        # Показываем заголовки
        headers = {}
        for col in range(1, min(ws.max_column + 1, 100)):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[col] = str(h).strip().replace('\n', ' ')[:40]
        log(f"  Заголовки: {headers}")

        # Показываем первую строку данных
        if ws.max_row >= 2:
            log(f"  Строка 2 (пример данных):")
            for col in range(1, min(ws.max_column + 1, 100)):
                v = ws.cell(row=2, column=col).value
                if v is not None:
                    log(f"    Col {col}: {repr(str(v)[:60])} (type: {type(v).__name__})")

        wb.close()
    except Exception as e:
        log(f"  ОШИБКА: {e}")

# ============================================================
# ШАГ 4: Поиск совпадений в файлах заказов
# ============================================================
log(f"\n{'=' * 60}")
log("ПОИСК СОВПАДЕНИЙ")
log("=" * 60)

# Строим обратный индекс для быстрого поиска
search_index = {}
for search_vals, template_row in search_values:
    for val in search_vals:
        if val not in search_index:
            search_index[val] = []
        search_index[val].append(template_row)

log(f"Поисковый индекс: {len(search_index)} уникальных значений")

updated_prices = {}  # template_row -> (price, filename, order_row)
processed_files = 0

for filename in sorted(os.listdir(orders_dir)):
    if not filename.endswith('.xlsx') or filename == 'шаблон.xlsx':
        continue
    file_path = os.path.join(orders_dir, filename)
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        processed_files += 1
        file_matches = 0

        for row_idx in range(2, ws.max_row + 1):
            # Собираем ВСЕ значения из строки для сравнения
            compare_values = set()
            for col_idx in range(1, min(ws.max_column + 1, 100)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    for bc in split_multiple_barcodes(cell_val):
                        # Добавляем только если выглядит как штрихкод/код (цифры)
                        cleaned = bc.strip()
                        if cleaned and (cleaned.isdigit() or (len(cleaned) > 4 and any(c.isdigit() for c in cleaned))):
                            compare_values.add(cleaned)

            # Ищем цену — берём из всех числовых столбцов, пробуем найти по заголовку
            # Сначала пробуем найти столбец с ценой по заголовку
            price = None
            price_col_candidates = []
            for col in range(1, min(ws.max_column + 1, 100)):
                h = ws.cell(row=1, column=col).value
                if h and any(word in str(h).lower() for word in ['цена', 'price', 'стоимость', 'сумма']):
                    price_col_candidates.append(col)

            # Если нашли столбцы с ценой по заголовку — берём первый с числовым значением
            for pc in price_col_candidates:
                val = ws.cell(row=row_idx, column=pc).value
                p = to_float(val)
                if p is not None and p > 0:
                    price = p
                    break

            # Если не нашли по заголовку — пробуем оригинальные столбцы (0-based: 75 = col BX)
            if price is None:
                for col_0based in [75]:
                    col_1based = col_0based + 1
                    if col_1based <= ws.max_column:
                        val = ws.cell(row=row_idx, column=col_1based).value
                        p = to_float(val)
                        if p is not None and p > 0:
                            price = p
                            break

            # Ищем совпадения
            for cv in compare_values:
                if cv in search_index:
                    for template_row in search_index[cv]:
                        if price is not None:
                            # Запоминаем: если уже есть — перезаписываем (последний файл побеждает)
                            updated_prices[template_row] = (price, filename, row_idx, cv)
                            file_matches += 1

        log(f"  {filename}: найдено {file_matches} совпадений")
        wb.close()
    except Exception as e:
        log(f"  ОШИБКА в файле {filename}: {e}")

log(f"\nОбработано файлов: {processed_files}")
log(f"Уникальных строк шаблона с найденными ценами: {len(updated_prices)}")

if updated_prices:
    log(f"\nПримеры найденных совпадений (первые 10):")
    for row, (price, fname, orow, match_val) in list(updated_prices.items())[:10]:
        log(f"  Шаблон строка {row}: цена={price}, из файла={fname}, строка={orow}, совпадение по={match_val}")
else:
    log("\n⚠️ СОВПАДЕНИЙ НЕ НАЙДЕНО!")
    log("Возможные причины:")
    log("  1. Штрихкоды/коды в файлах заказов не совпадают с шаблоном")
    log("  2. Формат данных в файлах заказов отличается")
    log("  Проверьте лог выше — раздел ДИАГНОСТИКА ФАЙЛОВ ЗАКАЗОВ")

# ============================================================
# ШАГ 5: Обновление цен и расчёт разницы
# ============================================================
log(f"\n{'=' * 60}")
log("ОБНОВЛЕНИЕ ЦЕН")
log("=" * 60)

updated_count = 0
diff_count = 0

for template_row, (new_price, src_file, src_row, match_val) in updated_prices.items():
    agreed_value = template_ws.cell(row=template_row, column=price_agreed_col).value
    agreed_num = to_float(agreed_value)

    # Записываем фактическую цену
    actual_cell = template_ws.cell(row=template_row, column=price_actual_col)
    actual_cell.value = new_price
    actual_cell.number_format = '0.00'

    # Рассчитываем разницу
    if agreed_num is not None and new_price is not None:
        difference = round(new_price - agreed_num, 2)
        diff_cell = template_ws.cell(row=template_row, column=price_diff_col)
        diff_cell.value = difference
        diff_cell.number_format = '+0.00;-0.00;0.00'

        if abs(difference) > 0.01:
            template_ws.cell(row=template_row, column=price_agreed_col).fill = YELLOW_FILL
            actual_cell.fill = YELLOW_FILL
            diff_cell.fill = YELLOW_FILL
            diff_count += 1
            log(f"  Строка {template_row}: согласованная={agreed_num}, фактическая={new_price}, разница={difference}")
    else:
        template_ws.cell(row=template_row, column=price_diff_col).value = "N/A"
        log(f"  Строка {template_row}: фактическая={new_price}, согласованная=N/A (не удалось прочитать)")

    updated_count += 1

# ============================================================
# ШАГ 6: Очистка мусорного столбца 1063 (если есть)
# ============================================================
if template_ws.max_column > 50:
    log(f"\n⚠️ Обнаружены мусорные столбцы (max_column={template_ws.max_column})")
    log(f"  Очищаем столбец 1063...")
    for row in range(1, template_ws.max_row + 1):
        template_ws.cell(row=row, column=1063).value = None

# ============================================================
# ШАГ 7: Сохранение
# ============================================================
try:
    template_wb.save(template_path)
    log(f"\n{'=' * 60}")
    log(f"ИТОГО:")
    log(f"  Обновлено цен: {updated_count}")
    log(f"  Строк с различием > 0.01: {diff_count}")
    log(f"  Файл сохранён: {template_path}")
    log("=" * 60)
except Exception as e:
    log(f"ОШИБКА при сохранении: {e}")

template_wb.close()